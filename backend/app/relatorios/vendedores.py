"""
Relatório: Vendedores — HSGrowth CRM

⚠️ **Isto é uma cópia fiel de `~/projetos/relatorios-hsgrowth/vendedores.py`.**
A régua de negócio é de lá e continua sendo — o Erick a escreveu, conferiu e o
Nicholson já aprovou o formato. Aqui mudaram exatamente duas coisas, para o
relatório poder ser gerado pelo agente em vez de na máquina de alguém:

  1. A conexão sai do cadastro de conectores do HS.OS (`integrations`), não do
     módulo `bancos`, que só existe no computador do Erick.
  2. `gerar()` devolve os **bytes** do arquivo em vez de gravar em `saidas/`.
     Quem grava é o endpoint, no diretório de uploads, para virar um link.

**Não melhore a régua aqui.** Se ela mudar, muda lá primeiro e esta cópia
acompanha — senão o número do agente e o número do Erick divergem em silêncio,
que é o defeito que este repositório passou a semana inteira consertando. (formato espelhado no de SDR, que o chefe gostou).

Foco: ESFORÇO (tarefas de cada vendedor) + o que está TRAVADO (cards parados que
deviam estar se movendo) + RESULTADO (negócios ganhos).

Métricas (janela padrão: 30 dias — ajustável via `python vendedores.py 7`):
- Atividades = tarefas (card_tasks) CONCLUÍDAS no período, por tipo (Ligação, Follow-up,
  Tarefa, Reunião, Outras).
- Negócios ganhos = cards fechados (is_done_stage) no período, qtd + valor.
Régua dos cards (snapshot do AGORA, sem janela de período): todo card aberto — nem
ganho (is_done_stage) nem perdido (is_lost_stage) — vinculado ao vendedor por
assigned_to_id cai em exatamente UM dos dois grupos, pelo último toque:
- 🔴 Parados = ZERO toque nos últimos 7 dias.
- 🔵 Em andamento = teve algum toque nos últimos 7 dias.
Toque = activities (inclui card_moved) OU card_notes OU card_tasks concluída; a
criação do card é o piso, então card novo nunca nasce parado.
Sem corte de idade: o corte antigo de 60 dias escondia justamente os cards mais
abandonados (caíam em "em andamento" como se estivessem saudáveis). A idade virou
coluna própria, pra distinguir "novo e travado" de card que já devia ir pra perdido.

Saída: Excel com Resumo + uma aba por vendedor (cards parados card a card, com link).
"""
import io
import os
from datetime import date

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ⚠️ O período vira parâmetro de `gerar()`. Estes são só os padrões — o original
# os lia de `sys.argv`, que não existe dentro de uma requisição.
DIAS_PADRAO = 30
DIAS_PARADO = int(os.environ.get("DIAS_PARADO", "7"))
EXCLUIR_VENDEDORES = ["Erick Santos"]

AZUL = PatternFill("solid", fgColor="DDEBF7")
VERDE = PatternFill("solid", fgColor="C6EFCE")
AMBAR = PatternFill("solid", fgColor="FFEB9C")
VERMELHO = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITULO_FONT = Font(bold=True, size=14)
SUB_FONT = Font(bold=True, size=11, color="1F4E78")
LINK_FONT = Font(color="0563C1", underline="single")
_lado = Side(style="thin", color="D9D9D9")
BORDA = Border(left=_lado, right=_lado, top=_lado, bottom=_lado)

TIPOS = ["CALL", "FOLLOW_UP", "TASK", "MEETING"]
TIPO_LABEL = {"CALL": "Ligação", "FOLLOW_UP": "Follow-up", "TASK": "Tarefa",
              "MEETING": "Reunião"}

# filtro de "vendedor de verdade"
_excl = "".join(f" AND u.name <> '{n}'" for n in EXCLUIR_VENDEDORES)
COND_VEND = f"r.name = 'salesperson' AND u.is_active AND u.name NOT ILIKE '%teste%'{_excl}"

SQL_ATIV = f"""
SELECT u.name AS vendedor, t.task_type::text AS tipo, count(*) AS qtd
FROM card_tasks t
JOIN users u ON u.id = t.assigned_to_id
JOIN roles r ON r.id = u.role_id
WHERE {COND_VEND}
  AND t.is_completed AND t.completed_at >= now() - interval '__DIAS__ days'
GROUP BY 1, 2;
"""

SQL_GANHOS = f"""
SELECT u.name AS vendedor, count(*) AS qtd, COALESCE(sum(c.value), 0) AS valor
FROM cards c
JOIN lists l ON l.id = c.list_id
JOIN users u ON u.id = c.assigned_to_id
JOIN roles r ON r.id = u.role_id
WHERE {COND_VEND} AND l.is_done_stage = true AND c.is_deleted = false
  AND c.closed_at >= now() - interval '__DIAS__ days'
GROUP BY 1;
"""

SQL_PARADOS = f"""
WITH base AS (
  SELECT c.id, c.title, c.value, c.assigned_to_id, c.created_at,
         l.name AS etapa, b.name AS pipeline
  FROM cards c
  JOIN lists l ON l.id = c.list_id
  JOIN boards b ON b.id = l.board_id
  JOIN users u ON u.id = c.assigned_to_id
  JOIN roles r ON r.id = u.role_id
  WHERE c.is_deleted = false
    AND l.is_done_stage = false AND l.is_lost_stage = false
    AND b.category = 'vendas' AND b.is_deleted = false
    AND {COND_VEND}
),
tocado AS (
  SELECT base.*,
    GREATEST(
      base.created_at,
      COALESCE((SELECT max(a.created_at) FROM activities a WHERE a.card_id = base.id), base.created_at),
      COALESCE((SELECT max(n.created_at::timestamp) FROM card_notes n WHERE n.card_id = base.id), base.created_at),
      COALESCE((SELECT max(t.completed_at) FROM card_tasks t WHERE t.card_id = base.id AND t.is_completed), base.created_at)
    ) AS ult_toque,
    (SELECT min(t.due_date) FROM card_tasks t
       WHERE t.card_id = base.id AND t.is_completed = false
         AND COALESCE(t.is_cancelled, false) = false
         AND t.due_date::date >= CURRENT_DATE) AS proxima_tarefa,
    (SELECT t.title FROM card_tasks t
       WHERE t.card_id = base.id AND t.is_completed = false
         AND COALESCE(t.is_cancelled, false) = false
         AND t.due_date::date >= CURRENT_DATE
       ORDER BY t.due_date LIMIT 1) AS proxima_tarefa_titulo
  FROM base
)
SELECT (SELECT u.name FROM users u WHERE u.id = tocado.assigned_to_id) AS vendedor,
  tocado.id AS id, title, pipeline, etapa, value,
  (CURRENT_DATE - ult_toque::date) AS dias_parado,
  (CURRENT_DATE - created_at::date) AS idade_card,
  proxima_tarefa, proxima_tarefa_titulo
FROM tocado
WHERE (CURRENT_DATE - ult_toque::date) >= {DIAS_PARADO}
ORDER BY vendedor, (proxima_tarefa IS NOT NULL), dias_parado DESC;
"""


SQL_AQUISICAO = f"""
WITH base AS (
  SELECT c.id, c.title, c.value, c.assigned_to_id, c.created_at, l.name AS etapa
  FROM cards c
  JOIN lists l ON l.id = c.list_id
  JOIN boards b ON b.id = l.board_id
  JOIN users u ON u.id = c.assigned_to_id
  JOIN roles r ON r.id = u.role_id
  WHERE c.is_deleted = false
    AND l.is_done_stage = false AND l.is_lost_stage = false
    AND b.board_type = 'acquisition' AND b.is_deleted = false
    AND {COND_VEND}
),
tocado AS (
  SELECT base.*,
    GREATEST(
      base.created_at,
      COALESCE((SELECT max(a.created_at) FROM activities a WHERE a.card_id = base.id), base.created_at),
      COALESCE((SELECT max(n.created_at::timestamp) FROM card_notes n WHERE n.card_id = base.id), base.created_at),
      COALESCE((SELECT max(t.completed_at) FROM card_tasks t WHERE t.card_id = base.id AND t.is_completed), base.created_at)
    ) AS ult_toque,
    (SELECT min(t.due_date) FROM card_tasks t
       WHERE t.card_id = base.id AND t.is_completed = false
         AND COALESCE(t.is_cancelled, false) = false
         AND t.due_date::date >= CURRENT_DATE) AS proxima_tarefa,
    (SELECT t.title FROM card_tasks t
       WHERE t.card_id = base.id AND t.is_completed = false
         AND COALESCE(t.is_cancelled, false) = false
         AND t.due_date::date >= CURRENT_DATE
       ORDER BY t.due_date LIMIT 1) AS proxima_tarefa_titulo
  FROM base
)
SELECT (SELECT u.name FROM users u WHERE u.id = tocado.assigned_to_id) AS vendedor,
  tocado.id AS id, title, etapa, value,
  (CURRENT_DATE - ult_toque::date) AS dias_sem_toque,
  proxima_tarefa, proxima_tarefa_titulo
FROM tocado
WHERE (CURRENT_DATE - ult_toque::date) < {DIAS_PARADO}
ORDER BY vendedor, (proxima_tarefa IS NULL), proxima_tarefa NULLS LAST, dias_sem_toque DESC;
"""


def fetch(cur, sql):
    cur.execute(sql)
    cols = [c.name for c in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def estilizar_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, larguras):
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def gerar(dsn: str, dias: int = DIAS_PADRAO) -> tuple[bytes, str, dict]:
    """Monta o Excel e devolve `(bytes, nome_do_arquivo, resumo)`.

    `dsn` é a conexão de leitura do HSGrowth, que o chamador tira do conector
    cadastrado. Nada aqui escreve no banco.
    """
    # ⚠️ O período entra aqui, não no import: os SQL de módulo carregam o
    # marcador `__DIAS__` porque a f-string deles já foi avaliada com as outras
    # constantes. Trocar por `%s` exigiria mexer na régua, e a regra é copiar
    # fiel — ver o cabeçalho.
    dia = str(int(dias))
    sql_ativ = SQL_ATIV.replace("__DIAS__", dia)
    sql_ganhos = SQL_GANHOS.replace("__DIAS__", dia)
    base_url = "https://hsgrowth.healthsafetytech.com"
    with psycopg.connect(dsn, connect_timeout=15) as conn, conn.cursor() as cur:
        ativ = fetch(cur, sql_ativ)
        ganhos = fetch(cur, sql_ganhos)
        parados = fetch(cur, SQL_PARADOS)
        aquisicao = fetch(cur, SQL_AQUISICAO)

    hoje = date.today().isoformat()
    vends = sorted({r["vendedor"] for r in ativ} |
                   {r["vendedor"] for r in ganhos} |
                   {r["vendedor"] for r in parados} |
                   {r["vendedor"] for r in aquisicao})

    ativ_por = {v: {} for v in vends}
    for r in ativ:
        ativ_por[r["vendedor"]][r["tipo"]] = r["qtd"]
    ganho_por = {v: {"qtd": 0, "valor": 0.0} for v in vends}
    for r in ganhos:
        ganho_por[r["vendedor"]] = {"qtd": r["qtd"], "valor": float(r["valor"] or 0)}
    parado_por = {v: [] for v in vends}
    for r in parados:
        parado_por[r["vendedor"]].append(r)
    # cards da Aquisição tocados nos últimos DIAS_PARADO dias — os grupos já saem
    # disjuntos do banco (>= vs < DIAS_PARADO), não precisa deduplicar aqui
    aquis_por = {v: [] for v in vends}
    for r in aquisicao:
        v = r["vendedor"]
        if v in aquis_por:
            aquis_por[v].append(r)

    def tipos_do(v):
        d = ativ_por[v]
        conhec = {t: d.get(t, 0) for t in TIPOS}
        outras = sum(q for k, q in d.items() if k not in TIPOS)
        total = sum(conhec.values()) + outras
        return conhec, outras, total

    wb = Workbook()

    # ---------- RESUMO ----------
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = f"Vendedores · esforço, travados e resultado · últimos {dias} dias"
    ws["A1"].font = TITULO_FONT
    ws["A2"] = (f"Atividades = tarefas concluídas no período · Ganhos = fechados no período · "
                f"Cards = snapshot de hoje (nem ganho nem perdido): parado = sem NENHUM toque "
                f"há ≥{DIAS_PARADO}d, em andamento = tocado nos últimos {DIAS_PARADO}d · gerado em {hoje}")
    ws["A2"].font = Font(italic=True, color="808080")
    cab = (["Vendedor", "Atividades (total)"] + [TIPO_LABEL[t] for t in TIPOS] +
           ["Outras", f"🔵 Em andamento ({DIAS_PARADO}d)", "🔴 Cards parados",
            "Negócios ganhos", "R$ ganhos"])
    ws.append([]); ws.append(cab)
    estilizar_header(ws, ws.max_row, len(cab))

    tot = [0] * (len(TIPOS) + 6)
    for v in sorted(vends, key=lambda x: -tipos_do(x)[2]):
        conhec, outras, total = tipos_do(v)
        n_par = len(parado_por[v])
        n_aq = len(aquis_por[v])
        g = ganho_por[v]
        linha = ([v, total] + [conhec[t] for t in TIPOS] +
                 [outras, n_aq, n_par, g["qtd"], g["valor"]])
        ws.append(linha)
        row = ws.max_row
        ws.cell(row, 2).fill = VERDE
        ws.cell(row, len(cab) - 3).fill = AZUL   # Aquisição ativos
        if n_par:
            ws.cell(row, len(cab) - 2).fill = VERMELHO
        ws.cell(row, len(cab)).number_format = '#,##0.00'
        vals = [total] + [conhec[t] for t in TIPOS] + [outras, n_aq, n_par, g["qtd"], g["valor"]]
        for i, x in enumerate(vals):
            tot[i] += x
    ws.append(["TOTAL"] + tot)
    for c in range(1, len(cab) + 1):
        ws.cell(ws.max_row, c).font = Font(bold=True)
    ws.cell(ws.max_row, len(cab)).number_format = '#,##0.00'
    auto_width(ws, [20, 16, 10, 11, 9, 9, 8, 16, 15, 16, 12])

    # ---------- UMA ABA POR VENDEDOR ----------
    cab_det = ["#", "Card", "Pipeline", "Etapa", "Valor (R$)", "Dias parado",
               "Idade do card", "Situação", "Próxima tarefa", "Link"]
    for v in sorted(vends, key=lambda x: -tipos_do(x)[2]):
        conhec, outras, total = tipos_do(v)
        g = ganho_por[v]
        nome_aba = v[:28].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(nome_aba)
        ws["A1"] = v
        ws["A1"].font = TITULO_FONT
        ws["A3"] = f"Atividades realizadas ({dias}d): {total}"
        ws["A3"].font = SUB_FONT
        det = "   ".join(f"{TIPO_LABEL[t]}: {conhec[t]}" for t in TIPOS if conhec[t])
        if outras:
            det += f"   Outras: {outras}"
        ws["A4"] = det
        ws["A5"] = (f"Negócios ganhos ({dias}d): {g['qtd']}  (R$ {g['valor']:,.2f})   "
                    f"·   🔴 Cards parados: {len(parado_por[v])}   "
                    f"·   Em andamento na Aquisição: {len(aquis_por[v])}")
        ws["A5"].font = SUB_FONT

        # --- Seção 1: atrasados ---
        ws.append([]); ws.append([])
        ws.append(["🔴 Cards parados que deviam estar se movendo:"])
        ws.cell(ws.max_row, 1).font = SUB_FONT
        ws.append(cab_det)
        estilizar_header(ws, ws.max_row, len(cab_det))
        for d in parado_por[v]:
            futuro = d["proxima_tarefa"] is not None
            situacao = "⚠️ Com tarefa futura" if futuro else "🔴 Sem nada agendado"
            url = f"{base_url}/cards/{d['id']}"
            ws.append([d["id"], d["title"], d["pipeline"], d["etapa"],
                       float(d["value"] or 0), d["dias_parado"], d["idade_card"],
                       situacao, d["proxima_tarefa_titulo"] or "", url])
            row = ws.max_row
            ws.cell(row, 5).number_format = '#,##0.00'
            if d["dias_parado"] >= 30:
                ws.cell(row, 6).fill = VERMELHO
            elif d["dias_parado"] >= 15:
                ws.cell(row, 6).fill = AMBAR
            if d["idade_card"] >= 90:
                ws.cell(row, 7).fill = VERMELHO
            elif d["idade_card"] >= 60:
                ws.cell(row, 7).fill = AMBAR
            ws.cell(row, 8).fill = AMBAR if futuro else VERMELHO
            lc = ws.cell(row, 10); lc.hyperlink = url; lc.font = LINK_FONT
            for c in range(1, len(cab_det) + 1):
                ws.cell(row, c).border = BORDA
        if not parado_por[v]:
            ws.append(["— Nenhum card parado. 👏"])

        # --- Seção 2: em andamento na Aquisição (para trabalhar) ---
        cab_aq = ["#", "Card", "Etapa", "Valor (R$)", "Dias s/ toque",
                  "Próxima tarefa", "Agendada p/", "Link"]
        ws.append([]); ws.append([])
        ws.append([f"🔵 Em andamento na Aquisição (com toque nos últimos {DIAS_PARADO} dias):"])
        ws.cell(ws.max_row, 1).font = SUB_FONT
        ws.append(cab_aq)
        estilizar_header(ws, ws.max_row, len(cab_aq))
        for d in aquis_por[v]:
            url = f"{base_url}/cards/{d['id']}"
            prox = str(d["proxima_tarefa"].date()) if d["proxima_tarefa"] else ""
            ws.append([d["id"], d["title"], d["etapa"], float(d["value"] or 0),
                       d["dias_sem_toque"], d["proxima_tarefa_titulo"] or "", prox, url])
            row = ws.max_row
            ws.cell(row, 4).number_format = '#,##0.00'
            if d["dias_sem_toque"] >= 15:
                ws.cell(row, 5).fill = AMBAR
            if not d["proxima_tarefa"]:
                ws.cell(row, 6).fill = AMBAR   # sem próxima tarefa = precisa agendar
            lc = ws.cell(row, 8); lc.hyperlink = url; lc.font = LINK_FONT
            for c in range(1, len(cab_aq) + 1):
                ws.cell(row, c).border = BORDA
        if not aquis_por[v]:
            ws.append(["— Nenhum card em andamento na Aquisição."])
        # larguras compartilhadas pelas duas seções: cada coluna usa a maior necessidade
        # (ex.: F = "Dias parado" em cima, "Próxima tarefa" embaixo)
        auto_width(ws, [7, 34, 22, 20, 13, 28, 14, 44, 28, 48])

    buf = io.BytesIO()
    wb.save(buf)
    tot_par = sum(len(parado_por[v]) for v in vends)
    resumo = {"vendedores": len(vends), "cards_parados": tot_par,
              "dias": dias, "data": hoje}
    return buf.getvalue(), f"vendedores_{dias}d_{hoje}.xlsx", resumo
